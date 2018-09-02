# -*- coding: utf-8 -*-
import openpyxl
from shutil import copyfile


class XLSXprocessor(object):
    def __init__(self, filename):
        self.filename = filename
        self.workbook = None
        self.worksheet = None

    def openFile(self):
        self.backup()
        self.workbook = openpyxl.load_workbook(self.filename)
        self.worksheet = self.workbook.active

    def __iter__(self):
        for row in self.worksheet.iter_rows(min_row=2):
            yield row

    def backup(self):
        copyfile(self.filename, self.filename + ".bak")

    def save(self):
        self.workbook.save(self.filename)


if __name__ == '__main__':
    xp = XLSXprocessor("C:\Users\charzewski\Downloads\charzoApkaLista.xlsx")
    xp.openFile()

    # print xp.worksheet
    # print
    # ws = xp.worksheet
    # #print ws[1:3]
    # a = xp.iterRows()
    # print a.next()
    #
    # for row in ws.iter_rows(min_row=2):
    #     tmp = row[0:4]
    #     vals = [i.value for i in tmp]
    #     print vals
    #
    for i in xp:
        print i
    xp.save()
